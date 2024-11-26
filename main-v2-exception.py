import os
import tkinter as tk
import sys
import fitz
from tkinter import filedialog
import re
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk

# Determine the base path for the application, considering both frozen and non-frozen states
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

image1 = os.path.join(base_path, 'dev-logo.png')


class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output file paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for UI

        # Get the screen width and height
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 980) // 2  # Adjust the window width as needed
        y_position = (screen_height - 500) // 2  # Adjust the window height as needed

        # Set the window geometry
        self.master.geometry(f"980x500+{x_position}+{y_position}")

        # Heading
        heading_label = tk.Label(self.master, text="Content Rationalizer", font=("Helvetica", 26, "bold"), bg="#1a1a2e",
                                 fg="white")
        heading_label.pack(pady=30)

        # Input File Location
        input_frame = tk.Frame(self.master)
        input_frame.pack(pady=35)

        input_label = tk.Label(input_frame, text="Input Folder ")
        input_label.pack(side=tk.LEFT)

        input_entry = tk.Entry(input_frame, textvariable=self.input_folder_path, width=50)
        input_entry.pack(side=tk.LEFT)

        input_button = tk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        input_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=35)

        output_label = tk.Label(output_frame, text="Output Folder ")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Compare Buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=10)

        # Button to perform PDF rationalization (comparison)
        compare_button = tk.Button(compare_frame, text="Rationalise", font=("Helvetica", 10, "bold"),
                                   command=self.compare_pdfs, width=20, height=2, bg="white")
        compare_button.pack(side=tk.LEFT)

        # Button to calculate percentage similarity between PDFs
        similarity_button = tk.Button(compare_frame, text="Percentage Match", font=("Helvetica", 10, "bold"),
                                      command=self.compare_similarity, width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=(15, 0))

        # Display the logo image on the right side
        original_image = Image.open(image1)
        self.photo = ImageTk.PhotoImage(original_image)
        self.image_label = tk.Label(self.master, image=self.photo)
        self.image_label.pack(side="right", padx=35)

    # Browsing utility to browse the input folder
    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_path.set(folder_path)

    # Browsing utility to browse the output folder
    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)

    # Extracts all the paragraphs from the given PDF file
    def extract_paragraphs_from_pdf(self, file_path):
        paragraphs = []
        try:
            with fitz.open(file_path) as pdf_document:
                for page_number in range(pdf_document.page_count):
                    page = pdf_document.load_page(page_number)
                    text = page.get_text("text")

                    # Normalize line breaks to a standard format
                    standardized_text = re.sub(r'\r\n|\n|\r', '\n', text)
                    # Split text into paragraphs based on multiple line breaks
                    page_paragraphs = re.split(r'[\n]{2,}(?=\S)', standardized_text)

                    # Add all paragraphs to the list
                    paragraphs.extend(page_paragraphs)

            # Remove empty paragraphs and strip leading/trailing whitespace
            paragraphs = [p.strip() for p in paragraphs if p.strip()]
        except Exception as e:
            print(f"Error extracting text from {file_path}: {str(e)}")

        return paragraphs

    # Compare paragraphs based on the selected operation (Rationalize or Percentage Match)
    def compare_paragraphs(self, pdf_paths, compare):
        if compare == "pdfcompare":
            # Set to hold all unique paragraphs across PDFs
            all_paragraphs = set()
            pdf_paragraphs = {}
            for pdf_path in pdf_paths:
                paragraphs = set(self.extract_paragraphs_from_pdf(pdf_path))
                all_paragraphs.update(paragraphs)
                pdf_paragraphs[pdf_path] = paragraphs

            # Sort paragraphs for consistent ordering
            all_paragraphs = sorted(list(all_paragraphs))
            if all_paragraphs and all_paragraphs[0] == '':
                all_paragraphs.remove('')

            # Create a presence matrix representing paragraph existence in each PDF
            matrix = []
            for pdf_path in pdf_paths:
                row = [os.path.basename(pdf_path)]
                for paragraph in all_paragraphs:
                    row.append(1 if paragraph in pdf_paragraphs[pdf_path] else 0)
                matrix.append(row)

            return all_paragraphs, matrix

        elif compare == "percentage_match":
            # Collect all paragraphs from all PDFs
            all_paragraphs = []
            for pdf_path in pdf_paths:
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                all_paragraphs.extend(paragraphs)

            return all_paragraphs

    # Rationalize Operation - Compare PDFs and generate an Excel report
    def compare_pdfs(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        # Validate if both input and output folders are selected
        if not input_folder or not output_folder:
            print("Please select both input and output folders.")
            return

        # Get all PDF files in the input folder
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            print("No PDF files found in the input folder.")
            return

        # Compare paragraphs in PDFs
        common_paragraphs, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")

        # Prepare paragraph labels
        para = ["Paragraph " + str(i + 1) for i in range(len(common_paragraphs))]

        # Create an Excel workbook to save the results
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(common_paragraphs)):
            sheet.append([para[i], common_paragraphs[i]])

        # Create a new sheet for the matrix representation
        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Pdf Results")
        new_sheet.append(header_row)

        # Append each row to the new sheet
        for row in matrix:
            new_sheet.append(row)

        # Save the workbook with a timestamp in the output folder
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        location = os.path.join(output_folder, f"resultoutput-{format_time}.xlsx")
        workbook.save(location)
        workbook.close()
        print(f"Comparison results saved to {location}")

    # Percentage Match Operation - Calculate similarity between paragraphs of PDFs
    def compare_similarity(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        # Validate if both input and output folders are selected
        if not input_folder or not output_folder:
            print("Please select both input and output folders.")
            return

        # Get all PDF files in the input folder
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            print("No PDF files found in the input folder.")
            return

        # Extract all paragraphs from PDFs
        all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")
        para = ["Paragraph " + str(i + 1) for i in range(len(all_paragraphs))]

        # Create an Excel workbook to save the percentage match results
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(all_paragraphs)):
            sheet.append([para[i], all_paragraphs[i]])

        # Create a new sheet for percentage similarity matrix
        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Percentage Match")
        new_sheet.append(header_row)
        matrix = []

        # Function to sort words in a paragraph for better comparison
        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        # Sort words in each paragraph for comparison
        sorted_paragraphs = [sort_words(x) for x in all_paragraphs]

        # Calculate similarity between each pair of paragraphs
        for para1 in range(0, len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(0, len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()
                temp_list.append(round(s * 100, 2))

            matrix.append(temp_list)

        # Append similarity matrix to the new sheet
        for row in range(len(matrix)):
            final_row = [para[row]] + matrix[row]
            new_sheet.append(final_row)

        # Save the workbook with a timestamp in the output folder
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        location = os.path.join(output_folder, f"percentage-{format_time}.xlsx")
        workbook.save(location)
        workbook.close()
        print(f"Percentage match results saved to {location}")


if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#1a1a2e")
    app = PDFComparerApp(root)
    root.mainloop()

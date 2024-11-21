import os
import tkinter as tk
import sys
import fitz
from tkinter import filedialog
import re
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk

if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

image1 = os.path.join(base_path, 'dev-logo.png')


class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output file paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for UI
        # Get the screen width and height
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 900) // 2  # Adjust the window width as needed
        y_position = (screen_height - 500) // 2  # Adjust the window height as needed

        # Set the window geometry
        self.master.geometry(f"900x500+{x_position}+{y_position}")

        # Heading
        heading_label = tk.Label(self.master, text="Content Rationalizer", font=("Helvetica", 26, "bold"), bg="#1a1a2e",
                                 fg="white")
        heading_label.pack(pady=30)

        # Input File Location
        input_frame = tk.Frame(self.master)
        input_frame.pack(pady=35)

        input_label = tk.Label(input_frame, text="Input Folder ")
        input_label.pack(side=tk.LEFT)

        input_entry = tk.Entry(input_frame, textvariable=self.input_folder_path, width=50)
        input_entry.pack(side=tk.LEFT)

        input_button = tk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        input_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=35)

        output_label = tk.Label(output_frame, text="Output Folder ")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Compare Buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=10)

        compare_button = tk.Button(compare_frame, text="Rationalise", font=("Helvetica", 10, "bold"),
                                   command=self.compare_pdfs, width=20, height=2, bg="white")
        compare_button.pack(side=tk.LEFT)

        similarity_button = tk.Button(compare_frame, text="Percentage Match", font=("Helvetica", 10, "bold"),
                                      command=self.compare_similarity, width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=(15, 0))

        if os.path.exists(image1):
            original_image = Image.open(image1)
            self.photo = ImageTk.PhotoImage(original_image)
            self.image_label = tk.Label(self.master, image=self.photo)
            self.image_label.pack(side="right", padx=35)
        else:
            print(f"Image file not found: {image1}")

    # Browsing utility to browse the Folders
    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        self.input_folder_path.set(folder_path)

    # Browsing utility to browse the Folders
    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        self.output_folder_path.set(folder_path)

    # Extracts all the paragraphs details from all the PDFs
    def extract_paragraphs_from_pdf(self, file_path):
        paragraphs = []
        try:
            with fitz.open(file_path) as pdf_document:
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    text = page.get_text("text")

                    standardized_text = re.sub(r'\r\n|\r|\n', '\n', text)
                    page_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', standardized_text)
                    paragraphs.extend(page_paragraphs)

                table_pattern = re.compile(r'\n[=]+\n')
                table_matches = table_pattern.finditer(text)
                for match in table_matches:
                    table_start = match.start()
                    table_end = match.end()
                    if table_end < len(text):
                        additional_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', text[table_end:].strip())
                        paragraphs.extend(additional_paragraphs)
                paragraphs.remove("")
        except Exception as e:
            print(f"Error extracting text from {file_path}: {str(e)}")
        return paragraphs

    # based on the button click Comparison or Percentage Match will be done
    def compare_paragraphs(self, pdf_paths, compare):
        if compare == "pdfcompare":
            # set holds unique pdf's
            all_paragraphs = set()
            # Extract paragraphs from all PDFs
            pdf_paragraphs = {}
            for pdf_path in pdf_paths:
                paragraphs = set(self.extract_paragraphs_from_pdf(pdf_path))
                all_paragraphs.update(paragraphs)
                pdf_paragraphs[pdf_path] = paragraphs

            all_paragraphs = sorted(list(all_paragraphs))
            print(all_paragraphs)
            print(len(all_paragraphs))
            if all_paragraphs[0] == '':
                all_paragraphs.remove('')

            # Create a matrix to represent the presence of paragraphs in each PDF
            matrix = []
            for pdf_path in pdf_paths:
                row = [pdf_path[30:]]
                for paragraph in all_paragraphs:
                    if paragraph != '':
                        row.append(1 if paragraph in pdf_paragraphs[pdf_path] else 0)
                matrix.append(row)
            return all_paragraphs, matrix
        elif compare == "percentage_match":
            all_paragraphs = []
            for pdf_path in pdf_paths:
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                print(len(paragraphs))
                all_paragraphs.extend(paragraphs)
            return all_paragraphs

    # Rationalize Operation
    def compare_pdfs(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        print(pdf_paths)

        common_paragraphs, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")
        print(len(common_paragraphs))

        para = []
        for i in range(len(common_paragraphs)):
            para.append("Paragraph " + str(i + 1))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(common_paragraphs)):
            sheet.append([para[i], common_paragraphs[i]])

        header_row = ["PDF"] + para
        print('\t'.join(header_row))
        new_sheet = workbook.create_sheet(title="Pdf Results")
        para.insert(0, "")
        new_sheet.append(para)
        for row in matrix:
            new_sheet.append(row)
        print("here")

        current_time = datetime.now()
        format_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        print("inn")
        location = output_folder + "\\percentage_" + re.sub(r'[^\w\s.-]', '', format_time) + ".xlsx"
        print(location)
        workbook.save(location)
        workbook.close()

    # Percentage Match Operation
    def compare_similarity(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")
        para = []
        for i in range(len(all_paragraphs)):
            para.append("Paragraph " + str(i + 1))

        workbook = openpyx1.Workbook()
        sheet = workbook.active
        for i in range(len(all_paragraphs)):
            sheet.append([para[i], all_paragraphs[i]])

        header_row = ["PDF"] + para
        print('\t'.join(header_row))
        new_sheet = workbook.create_sheet(title="Percentage Match")
        para.insert(0, "")
        new_sheet.append(para)
        matrix = []

        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        sorted_paragraphs = [sort_words(x) for x in all_paragraphs]

        for para1 in range(0, len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(0, len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()
                temp_list.append(round(s * 100, 2))
            matrix.append(temp_list)

        for i in range(len(all_paragraphs)):
            sheet.append([para[i], all_paragraphs[i]])

        header_row = ["PDF"] + para
        print('\t'.join(header_row))
        new_sheet = workbook.create_sheet(title="Percentage Match")
        para.insert(0, "")
        new_sheet.append(para)

        for row in matrix:
            new_sheet.append(row)
        print("here")

        current_time = datetime.now()
        format_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        print("inn")
        location = output_folder + "\\percentage_" + re.sub(r'[^\w\s.-]', '', format_time) + ".xlsx"
        print(location)
        workbook.save(location)
        workbook.close()


if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg='#1a1a2e')
    app = PDFComparerApp(root)
    root.mainloop()

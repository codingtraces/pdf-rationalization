import os
import tkinter as tk
import sys
import fitz  # PyMuPDF, used for PDF operations
from tkinter import filedialog
import re
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk
import threading  # For background threading
from tkinter import ttk  # For progress bar
import multiprocessing  # For parallel processing
from multiprocessing import cpu_count

# Determine the base path for resources
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS  # If the script is compiled, use the temporary directory
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # Otherwise, use the script directory

# Paths to the logo images
image1 = os.path.join(base_path, 'dev-logo.png')
image2 = os.path.join(base_path, 'dev-logo.png')

# Extract all paragraph details from a PDF file (moved outside the class for multiprocessing compatibility)
def extract_paragraphs_from_pdf(file_path):
    paragraphs = []
    try:
        # Open the PDF file using PyMuPDF
        with fitz.open(file_path) as pdf_document:
            for page_number in range(pdf_document.page_count):
                page = pdf_document.load_page(page_number)
                text = page.get_text("text")

                # Standardize line endings to '\n'
                standardized_text = re.sub(r'\r\n|\n|\r', '\n', text)
                # Split paragraphs based on double newlines
                page_paragraphs = re.split(r'[\n]{2,}(?=\S)', standardized_text)

                # Add all paragraphs from the PDF to the list
                paragraphs.extend(page_paragraphs)

                # Look for table patterns and extract additional paragraphs
                table_pattern = re.compile(r'\n[ ]+\n')
                table_matches = table_pattern.finditer(text)
                for match in table_matches:
                    table_start = match.start()
                    table_end = match.end()

                    # Extract additional paragraphs after the table
                    if table_end < len(text):
                        additional_paragraphs = re.split(r'[\n]{1,2}(?=\S)', text[table_end:].strip())
                        paragraphs.extend(additional_paragraphs)

        # Remove any empty strings from the list of paragraphs
        paragraphs = [p for p in paragraphs if p.strip()]
    except Exception as e:
        print(f"Error extracting text from {file_path}: {str(e)}")

    return paragraphs

class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output folder paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for the UI

        # Get the screen width and height to center the window
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 980) // 2  # Adjust the window width as needed
        y_position = (screen_height - 500) // 2  # Adjust the window height as needed

        # Set the window geometry
        self.master.geometry(f"980x500+{x_position}+{y_position}")

        # Load and display the logo images at the top left and top right
        try:
            if os.path.exists(image1) and os.path.exists(image2):
                original_image1 = Image.open(image1)
                original_image2 = Image.open(image2)

                # Resize images to a standard size (e.g., 100x100 pixels)
                resized_image1 = original_image1.resize((100, 100), Image.LANCZOS)
                resized_image2 = original_image2.resize((100, 100), Image.LANCZOS)

                self.photo1 = ImageTk.PhotoImage(resized_image1)
                self.photo2 = ImageTk.PhotoImage(resized_image2)

                # Create labels for the logos
                self.image_label1 = tk.Label(self.master, image=self.photo1, bg="#1a1a2e")
                self.image_label1.place(x=10, y=10, anchor="nw")

                self.image_label2 = tk.Label(self.master, image=self.photo2, bg="#1a1a2e")
                self.image_label2.place(x=870, y=10, anchor="ne")
            else:
                raise FileNotFoundError(f"One or both image files not found: {image1}, {image2}")
        except Exception as e:
            print(f"Error loading images: {str(e)}")

        # Heading label
        heading_label = tk.Label(self.master, text="Content Rationalizer", font=("Helvetica", 26, "bold"), bg="#1a1a2e",
                                 fg="white")
        heading_label.pack(pady=(120, 30))

        # Input File Location frame
        input_frame = tk.Frame(self.master)
        input_frame.pack(pady=35)

        input_label = tk.Label(input_frame, text="Input Folder ")
        input_label.pack(side=tk.LEFT)

        input_entry = tk.Entry(input_frame, textvariable=self.input_folder_path, width=50)
        input_entry.pack(side=tk.LEFT)

        input_button = tk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        input_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location frame
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=35)

        output_label = tk.Label(output_frame, text="Output Folder ")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Frame for compare buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=10)

        # Button for Rationalize operation
        compare_button = tk.Button(compare_frame, text="Rationalise", font=("Helvetica", 10, "bold"),
                                   command=self.start_compare_pdfs_thread, width=20, height=2, bg="white")
        compare_button.pack(side=tk.LEFT)

        # Button for Percentage Match operation
        similarity_button = tk.Button(compare_frame, text="Percentage Match", font=("Helvetica", 10, "bold"),
                                      command=self.start_compare_similarity_thread, width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=(15, 0))

        # Progress bar
        self.progress = ttk.Progressbar(self.master, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(pady=20)

    # Browse and select the input folder
    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_path.set(folder_path)

    # Browse and select the output folder
    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)

    # Start a thread for the Rationalize operation
    def start_compare_pdfs_thread(self):
        threading.Thread(target=self.compare_pdfs, daemon=True).start()

    # Start a thread for the Percentage Match operation
    def start_compare_similarity_thread(self):
        threading.Thread(target=self.compare_similarity, daemon=True).start()

    # Compare paragraphs between PDFs based on the specified comparison type
    def compare_paragraphs(self, pdf_paths, compare):
        with multiprocessing.Pool(processes=cpu_count()) as pool:
            all_paragraphs_list = pool.map(extract_paragraphs_from_pdf, pdf_paths)

        if compare == "pdfcompare":
            # Set to hold all unique paragraphs from all PDFs
            all_paragraphs = set()
            # Dictionary to store paragraphs from each PDF
            pdf_paragraphs = {}
            for pdf_path, paragraphs in zip(pdf_paths, all_paragraphs_list):
                paragraphs = set(paragraphs)
                all_paragraphs.update(paragraphs)
                pdf_paragraphs[pdf_path] = paragraphs

            # Sort paragraphs for consistent ordering
            all_paragraphs = sorted(list(all_paragraphs))

            # Create a matrix to represent the presence of paragraphs in each PDF
            matrix = []
            for pdf_path in pdf_paths:
                row = [os.path.basename(pdf_path)]  # Append PDF name (basename)
                for paragraph in all_paragraphs:
                    row.append(1 if paragraph in pdf_paragraphs[pdf_path] else 0)
                matrix.append(row)

            return all_paragraphs, matrix

        elif compare == "percentage_match":
            # Extract paragraphs from all PDFs into a single list
            all_paragraphs = []
            for paragraphs in all_paragraphs_list:
                all_paragraphs.extend(paragraphs)

            return all_paragraphs

    # Perform the Rationalize operation
    def compare_pdfs(self):
        self.update_progress(0)
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            print("Input and output folders must be selected.")
            self.update_progress(0)
            return

        # Get all PDF file paths from the input folder
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            print("No PDF files found in the input folder.")
            self.update_progress(0)
            return

        common_paragraphs, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")

        # Prepare paragraph labels
        para = ["Paragraph " + str(i + 1) for i in range(len(common_paragraphs))]

        # Create a new workbook and write the common paragraphs
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(common_paragraphs)):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', common_paragraphs[i])
            sheet.append([para[i], clean_paragraph])

        # Add matrix data to a new sheet
        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Pdf Results")
        new_sheet.append(header_row)

        for row in matrix:
            new_sheet.append(row)

        # Save the workbook with a timestamp
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        location = os.path.join(output_folder, f"resultoutput-{format_time}.xlsx")
        workbook.save(location)
        workbook.close()
        self.update_progress(100)

    # Perform the Percentage Match operation
    def compare_similarity(self):
        self.update_progress(0)
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            print("Input and output folders must be selected.")
            self.update_progress(0)
            return

        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            print("No PDF files found in the input folder.")
            self.update_progress(0)
            return

        all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")

        # Prepare paragraph labels
        para = ["Paragraph " + str(i + 1) for i in range(len(all_paragraphs))]

        # Create a new workbook and write the paragraphs
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(all_paragraphs)):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', all_paragraphs[i])
            sheet.append([para[i], clean_paragraph])

        # Add similarity data to a new sheet
        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Percentage Match")
        new_sheet.append(header_row)
        matrix = []

        # Helper function to sort words in a paragraph for comparison
        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        sorted_paragraphs = [sort_words(x) for x in all_paragraphs]

        # Calculate the similarity matrix
        for para1 in range(0, len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(0, len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()  # Get the similarity ratio
                temp_list.append(round(s * 100, 2))  # Convert to percentage and round

            matrix.append(temp_list)

        # Write the similarity matrix to the workbook
        for row in range(len(matrix)):
            final_row = [para[row]] + matrix[row]
            new_sheet.append(final_row)

        # Save the workbook with a timestamp
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        location = os.path.join(output_folder, f"percentage-{format_time}.xlsx")
        workbook.save(location)
        workbook.close()
        self.update_progress(100)

    # Update progress bar
    def update_progress(self, value):
        self.progress["value"] = value
        self.master.update_idletasks()


if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#1a1a2e")
    app = PDFComparerApp(root)
    root.mainloop()

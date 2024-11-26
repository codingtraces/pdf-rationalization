import os
import tkinter as tk
import sys
import fitz  # PyMuPDF, used for PDF operations
from tkinter import filedialog
import re
import openpyxl
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk
import threading
import logging
import time
import csv
import hashlib
from functools import lru_cache
from multiprocessing import Pool, cpu_count

# Configure logging for detailed debugging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Determine the base path for resources
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS  # If the script is compiled, use the temporary directory
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # Otherwise, use the script directory

# Paths to the logo images
image1 = os.path.join(base_path, 'dev-logo.png')
image2 = os.path.join(base_path, 'dev-logo.png')


def hash_paragraph(paragraph):
    return hashlib.sha256(paragraph.encode('utf-8')).hexdigest()


@lru_cache(maxsize=1024)
def extract_paragraphs_from_pdf_cached(file_path, file_modified_time, min_char_count):
    paragraphs = []
    try:
        logging.info(f"Extracting text from {file_path} using PyMuPDF.")
        with fitz.open(file_path) as doc:
            text = ""
            for page in doc:
                text += page.get_text()

        lines = text.splitlines()
        paragraph = ""
        for line in lines:
            if line.strip():
                if paragraph:
                    paragraph += " " + line.strip()
                else:
                    paragraph = line.strip()
            else:
                if paragraph:
                    if len(paragraph) >= min_char_count:
                        paragraphs.append(paragraph.strip())
                    paragraph = ""
        if paragraph and len(paragraph) >= min_char_count:
            paragraphs.append(paragraph.strip())

        combined_paragraphs = []
        temp_paragraph = ""
        for para in paragraphs:
            if len(para.split()) < 20:
                temp_paragraph += " " + para
            else:
                if temp_paragraph:
                    combined_paragraphs.append(temp_paragraph.strip())
                    temp_paragraph = ""
                combined_paragraphs.append(para)
        if temp_paragraph:
            combined_paragraphs.append(temp_paragraph.strip())

        return combined_paragraphs
    except Exception as e:
        logging.error(f"Error extracting text from {file_path}: {str(e)}")
    return paragraphs


def extract_paragraphs_from_pdf(file_path, min_char_count):
    file_modified_time = os.path.getmtime(file_path)
    return extract_paragraphs_from_pdf_cached(file_path, file_modified_time, min_char_count)


def process_pdfs_in_parallel(pdf_paths, min_char_count):
    with Pool(processes=cpu_count()) as pool:
        return pool.starmap(extract_paragraphs_from_pdf, [(pdf_path, min_char_count) for pdf_path in pdf_paths])


def generate_common_hashes_and_matrix(pdf_paths, all_paragraphs):
    all_hashes = set()
    pdf_hashes = {}
    for index, paragraphs in enumerate(all_paragraphs):
        all_hashes.update(paragraphs)
        pdf_hashes[pdf_paths[index]] = paragraphs

    all_hashes = sorted(list(all_hashes))
    matrix = [
        [os.path.basename(pdf_path)] + [1 if hash_value in pdf_hashes[pdf_path] else 0 for hash_value in all_hashes]
        for pdf_path in pdf_paths
    ]
    return all_hashes, matrix


def filter_matrix_and_hashes(common_hashes, matrix):
    filtered_hashes = []
    filtered_matrix = []

    for i, hash_value in enumerate(common_hashes):
        if any(row[i + 1] == 1 for row in matrix) and not all(row[i + 1] == 1 for row in matrix):
            filtered_hashes.append(hash_value)

    for row in matrix:
        filtered_row = [row[0]] + [row[i + 1] for i, hash_value in enumerate(common_hashes) if hash_value in filtered_hashes]
        if any(filtered_row[1:]):
            filtered_matrix.append(filtered_row)

    return filtered_hashes, filtered_matrix


def write_results(output_folder, filename_prefix, common_hashes, matrix, file_type):
    common_hashes, matrix = filter_matrix_and_hashes(common_hashes, matrix)

    current_time = datetime.now()
    format_time = current_time.strftime("%Y%m%d%H%M%S")
    if file_type == "excel":
        output_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.xlsx")
        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Common Paragraphs")
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(common_hashes):
            sheet.append([f"Paragraph {i + 1}", paragraph])

        new_sheet = workbook.create_sheet(title="Matrix")
        header_row = ["PDF"] + [f"Paragraph {i + 1}" for i in range(len(common_hashes))]
        new_sheet.append(header_row)
        for row in matrix:
            new_sheet.append(row)

        workbook.save(output_file)
        workbook.close()
        logging.info(f"Results are saved in the file: {output_file}")
    elif file_type == "csv":
        output_csv_common = os.path.join(output_folder, f"{filename_prefix}_common_{format_time}.csv")
        output_csv_matrix = os.path.join(output_folder, f"{filename_prefix}_matrix_{format_time}.csv")

        # Write Common Paragraphs to CSV
        with open(output_csv_common, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Paragraph ID", "Hash"])
            for i, hash_value in enumerate(common_hashes):
                writer.writerow([f"Paragraph {i + 1}", hash_value])

        # Write Matrix to CSV
        with open(output_csv_matrix, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            header_row = ["PDF"] + [f"Paragraph {i + 1}" for i in range(len(common_hashes))]
            writer.writerow(header_row)
            for row in matrix:
                writer.writerow(row)

        logging.info(f"CSV Results are saved in the files: {output_csv_common}, {output_csv_matrix}")
    elif file_type == "html":
        output_html_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.html")
        with open(output_html_file, 'w', encoding='utf-8') as file:
            file.write("<html><head><title>Rationalized Result</title></head><body>")
            file.write("<h1>Common Paragraphs</h1>")
            file.write("<table border='1'><tr><th>Paragraph ID</th><th>Content</th></tr>")
            for i, paragraph in enumerate(common_hashes):
                file.write(f"<tr><td>Paragraph {i + 1}</td><td>{paragraph}</td></tr>")
            file.write("</table>")

            file.write("<h1>Matrix</h1>")
            file.write("<table border='1'><tr><th>PDF</th>")
            for i in range(len(common_hashes)):
                file.write(f"<th>Paragraph {i + 1}</th>")
            file.write("</tr>")
            for row in matrix:
                file.write("<tr>" + "".join([f"<td>{cell}</td>" for cell in row]) + "</tr>")
            file.write("</table>")
            file.write("</body></html>")

        logging.info(f"HTML Results are saved in the file: {output_html_file}")


def calculate_similarity_matrix(paragraphs):
    def sort_words(paragraph):
        return ' '.join(sorted(paragraph.split()))

    sorted_paragraphs = [sort_words(x) for x in paragraphs]
    matrix = []
    for para1 in range(0, len(sorted_paragraphs)):
        temp_list = []
        for para2 in range(0, len(sorted_paragraphs)):
            m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
            s = m.ratio()
            temp_list.append(round(s * 100, 2))
        matrix.append(temp_list)
    return matrix



def write_similarity_html(output_folder, filename_prefix, all_paragraphs):
    current_time = datetime.now()
    format_time = current_time.strftime("%Y%m%d%H%M%S")
    output_html_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.html")

    with open(output_html_file, 'w', encoding='utf-8') as file:
        file.write("<html><head><title>Similarity Report</title></head><body>")
        file.write("<h1>Paragraphs</h1>")
        file.write("<table border='1'><tr><th>Paragraph ID</th><th>Content</th></tr>")
        for i, paragraph in enumerate(all_paragraphs):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
            file.write(f"<tr><td>Paragraph {i + 1}</td><td>{clean_paragraph}</td></tr>")
        file.write("</table>")

        file.write("<h1>Similarity Matrix (Above 90%)</h1>")
        file.write("<table border='1'><tr><th>Paragraph</th>")
        for i in range(len(all_paragraphs)):
            file.write(f"<th>Paragraph {i + 1}</th>")
        file.write("</tr>")
        matrix = calculate_similarity_matrix(all_paragraphs)
        for row_idx, row in enumerate(matrix):
            filtered_row = [f"<td>{cell}</td>" if cell >= 90 else "<td></td>" for cell in row]
            if any(cell >= 90 for cell in row):
                file.write(f"<tr><td>Paragraph {row_idx + 1}</td>" + "".join(filtered_row) + "</tr>")
        file.write("</table>")
        file.write("</body></html>")

    logging.info(f"HTML Results are saved in the file: {output_html_file}")


class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output folder paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()
        self.min_char_count = tk.IntVar(value=100)  # Default minimum character count

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for the UI
        self.configure_window()
        self.create_heading_frame()
        self.create_input_output_frames()
        self.create_min_char_count_frame()
        self.create_compare_buttons()

    def configure_window(self):
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x_position = (screen_width - 980) // 2
        y_position = (screen_height - 600) // 2
        self.master.geometry(f"980x600+{x_position}+{y_position}")

    def create_heading_frame(self):
        heading_frame = tk.Frame(self.master, bg="#1a1a2e")
        heading_frame.pack(fill=tk.X, pady=10, padx=10)

        self.load_image(heading_frame, image1, "left")
        heading_label = self.create_label(heading_frame, "Content Rationalizer", font=("Helvetica", 26, "bold"),
                                          bg="#1a1a2e", fg="white")
        heading_label.pack(side="left", expand=True)
        self.load_image(heading_frame, image2, "right")

    def load_image(self, frame, image_path, side):
        try:
            if os.path.exists(image_path):
                original_image = Image.open(image_path).resize((80, 80), Image.LANCZOS)
                photo = ImageTk.PhotoImage(original_image)
                image_label = tk.Label(frame, image=photo, bg="#1a1a2e")
                image_label.image = photo  # Keep reference to avoid garbage collection
                image_label.pack(side=side, padx=10)
            else:
                raise FileNotFoundError(f"Image file not found: {image_path}")
        except Exception as e:
            logging.error(f"Error loading image: {str(e)}")

    def create_input_output_frames(self):
        self.create_folder_frame("Input Folder ", self.input_folder_path, self.browse_input_folder)
        self.create_folder_frame("Output Folder ", self.output_folder_path, self.browse_output_folder)

    def create_min_char_count_frame(self):
        frame = tk.Frame(self.master)
        frame.pack(pady=10)
        self.create_label(frame, "Minimum Characters for Rationalization: ", font=("Helvetica", 12)).pack(side=tk.LEFT)
        self.create_entry(frame, textvariable=self.min_char_count, width=10).pack(side=tk.LEFT, padx=(5, 0))

    def create_folder_frame(self, label_text, path_variable, browse_command):
        frame = tk.Frame(self.master)
        frame.pack(pady=10)
        self.create_label(frame, label_text, font=("Helvetica", 12)).pack(side=tk.LEFT)
        self.create_entry(frame, textvariable=path_variable, width=50).pack(side=tk.LEFT, padx=(5, 0))
        self.create_button(frame, "Browse", browse_command, font=("Helvetica", 10), width=10).pack(side=tk.LEFT,
                                                                                                   padx=(10, 0))

    def create_compare_buttons(self):
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=20, padx=10, fill=tk.X)

        button_texts = [
            "Rationalise (Excel)",
            "Rationalise (CSV)",
            "Rationalise (HTML)",
            "Percentage Match (Excel)",
            "Percentage Match (CSV)",
            "Percentage Match (HTML)"
        ]

        button_commands = [
            self.compare_pdfs_excel,
            self.compare_pdfs_csv,
            self.compare_pdfs_html,
            self.compare_similarity_excel,
            self.compare_similarity_csv,
            self.compare_similarity_html
        ]

        for i in range(len(button_texts)):
            button = self.create_button(compare_frame, button_texts[i],
                                        lambda cmd=button_commands[i]: threading.Thread(target=cmd).start(),
                                        font=("Helvetica", 10, "bold"), width=25, height=2, bg="white")
            button.grid(row=i // 3, column=i % 3, padx=10, pady=10, sticky='nsew')

        for i in range(3):
            compare_frame.grid_columnconfigure(i, weight=1)

    def create_label(self, frame, text, **kwargs):
        return tk.Label(frame, text=text, **kwargs)

    def create_entry(self, frame, textvariable, **kwargs):
        return tk.Entry(frame, textvariable=textvariable, **kwargs)

    def create_button(self, frame, text, command, **kwargs):
        return tk.Button(frame, text=text, command=command, **kwargs)

    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_path.set(folder_path)

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)

    def compare_pdfs_excel(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            common_hashes, matrix = generate_common_hashes_and_matrix(pdf_paths, all_paragraphs)
            write_results(output_folder, "rationalized_result", common_hashes, matrix, "excel")
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_pdfs_csv(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            common_hashes, matrix = generate_common_hashes_and_matrix(pdf_paths, all_paragraphs)
            write_results(output_folder, "rationalized_result", common_hashes, matrix, "csv")
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_pdfs_html(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            common_hashes, matrix = generate_common_hashes_and_matrix(pdf_paths, all_paragraphs)
            write_results(output_folder, "rationalized_result", common_hashes, matrix, "html")
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_similarity_excel(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            combined_paragraphs = [para for sublist in all_paragraphs for para in sublist]
            self.save_similarity_excel(output_folder, "percentage_report", combined_paragraphs)
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_similarity_csv(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            combined_paragraphs = [para for sublist in all_paragraphs for para in sublist]
            self.save_similarity_csv(output_folder, "percentage_report", combined_paragraphs)
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_similarity_html(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            min_char_count = self.min_char_count.get()
            all_paragraphs = process_pdfs_in_parallel(pdf_paths, min_char_count)
            combined_paragraphs = [para for sublist in all_paragraphs for para in sublist]
            write_similarity_html(output_folder, "percentage_report", combined_paragraphs)
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")

        self.log_processing_time(start_time)

    def get_input_output_paths(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            logging.error("Input and output folders must be selected.")
            return None, None, None

        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            logging.error("No PDF files found in the input folder.")
            return None, None, None

        return input_folder, output_folder, pdf_paths

    def save_similarity_excel(self, output_folder, filename_prefix, all_paragraphs):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.xlsx")

        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Paragraphs")
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(all_paragraphs):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
            sheet.append([f"Paragraph {i + 1}", clean_paragraph])

        new_sheet = workbook.create_sheet(title="Similarity Matrix (Above 90%)")
        header_row = ["Paragraph"] + [f"Paragraph {i + 1}" for i in range(len(all_paragraphs))]
        new_sheet.append(header_row)
        matrix = calculate_similarity_matrix(all_paragraphs)

        for row_idx, row in enumerate(matrix):
            if any(cell >= 90 for cell in row):
                filtered_row = [cell if cell >= 90 else None for cell in row]
                final_row = [f"Paragraph {row_idx + 1}"] + filtered_row
                new_sheet.append(final_row)

        workbook.save(output_file)
        workbook.close()
        logging.info(f"Results are saved in the file: {output_file}")

    def save_similarity_csv(self, output_folder, filename_prefix, all_paragraphs):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_csv_paragraphs = os.path.join(output_folder, f"{filename_prefix}_paragraphs_{format_time}.csv")
        output_csv_matrix = os.path.join(output_folder, f"{filename_prefix}_matrix_{format_time}.csv")

        # Write Paragraphs to CSV
        with open(output_csv_paragraphs, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Paragraph ID", "Content"])
            for i, paragraph in enumerate(all_paragraphs):
                clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
                writer.writerow([f"Paragraph {i + 1}", clean_paragraph])

        # Write Similarity Matrix to CSV
        matrix = calculate_similarity_matrix(all_paragraphs)
        with open(output_csv_matrix, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            header_row = ["Paragraph"] + [f"Paragraph {i + 1}" for i in range(len(all_paragraphs))]
            writer.writerow(header_row)
            for row_idx, row in enumerate(matrix):
                if any(cell >= 90 for cell in row):
                    filtered_row = [cell if cell >= 90 else '' for cell in row]
                    writer.writerow([f"Paragraph {row_idx + 1}"] + filtered_row)

        logging.info(f"CSV Results are saved in the files: {output_csv_paragraphs}, {output_csv_matrix}")

    def log_processing_time(self, start_time):
        end_time = time.time()
        elapsed_time = end_time - start_time
        logging.info(f"Processing completed in {elapsed_time:.2f} seconds.")



if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#1a1a2e")
    app = PDFComparerApp(root)
    root.mainloop()

import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import fitz
import re
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk
import string
import threading

# Determine the base path for image resources based on whether the script is frozen (e.g., packed as an exe)
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

image1 = os.path.join(base_path, 'dev-logo.png')
image2 = os.path.join(base_path, 'dev-logo.png')

class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output file paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Get the screen width and height
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 900) // 2
        y_position = (screen_height - 500) // 2

        # Set the window geometry
        self.master.geometry(f"900x500+{x_position}+{y_position}")

        # Display two images on the top left and top right of the GUI
        original_image1 = Image.open(image1)
        original_image2 = Image.open(image2)
        resized_image1 = original_image1.resize((original_image1.width // 2, original_image1.height // 2), Image.LANCZOS)
        resized_image2 = original_image2.resize((original_image2.width // 2, original_image2.height // 2), Image.LANCZOS)
        self.photo1 = ImageTk.PhotoImage(resized_image1)
        self.photo2 = ImageTk.PhotoImage(resized_image2)

        self.image_label1 = tk.Label(self.master, image=self.photo1, bg='#1a1a2e')
        self.image_label1.place(x=10, y=10, anchor='nw')

        self.image_label2 = tk.Label(self.master, image=self.photo2, bg='#1a1a2e')
        self.image_label2.place(x=880, y=10, anchor='ne')

        # Heading
        heading_label = tk.Label(self.master, text="Content Rationalizer", font=("Helvetica", 26, "bold"), bg='#1a1a2e', fg='white')
        heading_label.pack(pady=30)

        # Input File Location
        input_frame = tk.Frame(self.master)
        input_frame.pack(pady=15)

        input_label = tk.Label(input_frame, text="Input Folder")
        input_label.pack(side=tk.LEFT)

        input_entry = tk.Entry(input_frame, textvariable=self.input_folder_path, width=50)
        input_entry.pack(side=tk.LEFT)

        input_button = tk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        input_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=15)

        output_label = tk.Label(output_frame, text="Output Folder ")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Compare Buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=20)

        compare_button = tk.Button(compare_frame, text="Rationalize", font=("Helvetica", 10, "bold"), command=lambda: threading.Thread(target=self.compare_pdfs).start(), width=20, height=2, bg='white')
        compare_button.pack(side=tk.LEFT)

        similarity_button = tk.Button(compare_frame, text="Percentage Match", font=("Helvetica", 10, "bold"), command=lambda: threading.Thread(target=self.compare_similarity).start(), width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=(15, 0))

    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        self.input_folder_path.set(folder_path)

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        self.output_folder_path.set(folder_path)

    def show_progress(self, message):
        messagebox.showinfo("Progress", message)

    def extract_paragraphs_from_pdf(self, file_path):
        paragraphs = []
        try:
            with fitz.open(file_path) as pdf_document:
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    text = page.get_text("text")
                    standardized_text = re.sub(r'\r\n|\r|\n', '\n', text)
                    page_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', standardized_text)
                    paragraphs.extend(page_paragraphs)

                    table_pattern = re.compile(r'\n[=]+\n')
                    table_matches = table_pattern.finditer(text)
                    for match in table_matches:
                        table_start = match.start()
                        table_end = match.end()

                        if table_end < len(text):
                            additional_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', text[table_end:].strip())
                            paragraphs.extend(additional_paragraphs)

                    paragraphs = [p for p in paragraphs if p.strip() != ""]

        except Exception as e:
            messagebox.showerror("Error", f"Error extracting text from {file_path}: {str(e)}")

        return paragraphs

    def remove_illegal_characters(self, value):
        if isinstance(value, str):
            return ''.join(c for c in value if c in string.printable and c not in '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F')
        return value

    def compare_paragraphs(self, pdf_paths, compare):
        if compare == "pdfcompare":
            all_paragraphs = set()
            pdf_paragraphs = {}
            for pdf_path in pdf_paths:
                paragraphs = set(self.extract_paragraphs_from_pdf(pdf_path))
                all_paragraphs.update(paragraphs)
                pdf_paragraphs[pdf_path] = paragraphs

            all_paragraphs = sorted(list(all_paragraphs))
            if all_paragraphs and all_paragraphs[0] == '':
                all_paragraphs.remove('')

            matrix = []
            for pdf_path in pdf_paths:
                row = [os.path.basename(pdf_path)]
                for paragraph in all_paragraphs:
                    row.append(1 if paragraph in pdf_paragraphs[pdf_path] else 0)
                matrix.append(row)

            return all_paragraphs, matrix

        elif compare == "percentage_match":
            all_paragraphs = []
            for pdf_path in pdf_paths:
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                all_paragraphs.extend(paragraphs)

            return all_paragraphs

    def compare_pdfs(self):
        self.show_progress("Rationalization in progress...")

        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]

        common_paragraphs, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")

        para = ["Paragraph " + str(i + 1) for i in range(len(common_paragraphs))]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(common_paragraphs)):
            sanitized_paragraph = self.remove_illegal_characters(common_paragraphs[i])
            sheet.append([para[i], sanitized_paragraph])

        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Pdf Results")
        new_sheet.append(header_row)

        for row in matrix:
            new_sheet.append(row)

        current_time = datetime.now()
        format_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        location = os.path.join(output_folder, f"output_{format_time}.xlsx")
        workbook.save(location)
        workbook.close()

        self.show_progress("Rationalization completed successfully!")

    def compare_similarity(self):
        self.show_progress("Percentage match analysis in progress...")

        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()
        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")

        para = ["Paragraph " + str(i + 1) for i in range(len(all_paragraphs))]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(len(all_paragraphs)):
            sanitized_paragraph = self.remove_illegal_characters(all_paragraphs[i])
            sheet.append([para[i], sanitized_paragraph])

        header_row = ["PDF"] + para
        new_sheet = workbook.create_sheet(title="Percentage Match")
        new_sheet.append(header_row)

        matrix = []

        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        sorted_paragraphs = [sort_words(x) for x in all_paragraphs]

        for para1 in range(len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()
                temp_list.append(round(s * 100, 2))

            matrix.append(temp_list)

        for row in range(len(matrix)):
            final_row = [para[row]] + matrix[row]
            new_sheet.append(final_row)

        current_time = datetime.now()
        format_time = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        location = os.path.join(output_folder, f"percentage_{format_time}.xlsx")
        workbook.save(location)
        workbook.close()

        self.show_progress("Percentage match analysis completed successfully!")

if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg='#1a1a2e')
    app = PDFComparerApp(root)
    root.mainloop()
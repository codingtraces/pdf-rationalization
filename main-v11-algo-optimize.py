import os
import tkinter as tk
import sys
import fitz  # PyMuPDF, used for PDF operations
from tkinter import filedialog
import re
import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk
import threading
import logging
import time

# Configure logging for detailed debugging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Determine the base path for resources
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS  # If the script is compiled, use the temporary directory
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # Otherwise, use the script directory

# Paths to the logo images
image1 = os.path.join(base_path, 'dev-logo.png')
image2 = os.path.join(base_path, 'dev-logo.png')

class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output folder paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for the UI
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 980) // 2
        y_position = (screen_height - 500) // 2

        # Set the window geometry
        self.master.geometry(f"980x500+{x_position}+{y_position}")

        # Heading frame for logos and title
        heading_frame = tk.Frame(self.master, bg="#1a1a2e")
        heading_frame.pack(fill=tk.X, pady=10, padx=10)

        # Load and display the first logo image on the left
        try:
            if os.path.exists(image1):
                original_image1 = Image.open(image1).resize((100, 100), Image.LANCZOS)
                self.photo1 = ImageTk.PhotoImage(original_image1)
                self.image_label1 = tk.Label(heading_frame, image=self.photo1, bg="#1a1a2e")
                self.image_label1.pack(side="left", padx=10)
            else:
                raise FileNotFoundError(f"Image file not found: {image1}")
        except Exception as e:
            logging.error(f"Error loading image1: {str(e)}")

        # Heading label in the center
        heading_label = tk.Label(heading_frame, text="Content Rationalizer", font=("Helvetica", 26, "bold"),
                                 bg="#1a1a2e", fg="white")
        heading_label.pack(side="left", expand=True)

        # Load and display the second logo image on the right
        try:
            if os.path.exists(image2):
                original_image2 = Image.open(image2).resize((100, 100), Image.LANCZOS)
                self.photo2 = ImageTk.PhotoImage(original_image2)
                self.image_label2 = tk.Label(heading_frame, image=self.photo2, bg="#1a1a2e")
                self.image_label2.pack(side="right", padx=10)
            else:
                raise FileNotFoundError(f"Image file not found: {image2}")
        except Exception as e:
            logging.error(f"Error loading image2: {str(e)}")

        # Input File Location frame
        input_frame = tk.Frame(self.master)
        input_frame.pack(pady=20)

        input_label = tk.Label(input_frame, text="Input Folder ")
        input_label.pack(side=tk.LEFT)

        input_entry = tk.Entry(input_frame, textvariable=self.input_folder_path, width=50)
        input_entry.pack(side=tk.LEFT)

        input_button = tk.Button(input_frame, text="Browse", command=self.browse_input_folder)
        input_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location frame
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=20)

        output_label = tk.Label(output_frame, text="Output Folder ")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Frame for compare buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=10)

        # Button for Rationalize operation
        compare_button = tk.Button(compare_frame, text="Rationalise", font=("Helvetica", 10, "bold"),
                                   command=lambda: threading.Thread(target=self.compare_pdfs).start(), width=20,
                                   height=2, bg="white")
        compare_button.pack(side=tk.LEFT)

        # Button for Percentage Match operation
        similarity_button = tk.Button(compare_frame, text="Percentage Match", font=("Helvetica", 10, "bold"),
                                      command=lambda: threading.Thread(target=self.compare_similarity).start(),
                                      width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=(15, 0))

    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_path.set(folder_path)

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)

    def extract_paragraphs_from_pdf(self, file_path):
        paragraphs = []
        try:
            logging.info(f"Extracting text from {file_path} using PyMuPDF.")
            with fitz.open(file_path) as doc:
                text = ""
                for page in doc:
                    text += page.get_text()

            lines = text.splitlines()
            paragraph = ""
            for line in lines:
                if line.strip():
                    if paragraph:
                        paragraph += " " + line.strip()
                    else:
                        paragraph = line.strip()
                else:
                    if paragraph:
                        paragraphs.append(paragraph.strip())
                        paragraph = ""
            if paragraph:
                paragraphs.append(paragraph.strip())

            combined_paragraphs = []
            temp_paragraph = ""
            for para in paragraphs:
                if len(para.split()) < 20:
                    temp_paragraph += " " + para
                else:
                    if temp_paragraph:
                        combined_paragraphs.append(temp_paragraph.strip())
                        temp_paragraph = ""
                    combined_paragraphs.append(para)
            if temp_paragraph:
                combined_paragraphs.append(temp_paragraph.strip())

            return combined_paragraphs
        except Exception as e:
            logging.error(f"Error extracting text from {file_path}: {str(e)}")
        return paragraphs

    def compare_paragraphs(self, pdf_paths, compare):
        if compare == "pdfcompare":
            all_paragraphs = set()
            pdf_paragraphs = {}
            for index, pdf_path in enumerate(pdf_paths):
                logging.info(f"Extracting paragraphs from {pdf_path} ({index + 1}/{len(pdf_paths)})")
                paragraphs = set(self.extract_paragraphs_from_pdf(pdf_path))
                all_paragraphs.update(paragraphs)
                pdf_paragraphs[pdf_path] = paragraphs

            all_paragraphs = sorted(list(all_paragraphs))

            matrix = []
            for pdf_path in pdf_paths:
                row = [os.path.basename(pdf_path)]
                for paragraph in all_paragraphs:
                    row.append(1 if paragraph in pdf_paragraphs[pdf_path] else 0)
                matrix.append(row)

            return all_paragraphs, matrix

        elif compare == "percentage_match":
            all_paragraphs = []
            for index, pdf_path in enumerate(pdf_paths):
                logging.info(f"Extracting paragraphs from {pdf_path} ({index + 1}/{len(pdf_paths)})")
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                all_paragraphs.extend(paragraphs)

            return all_paragraphs

    def compare_pdfs(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            logging.error("Input and output folders must be selected.")
            return

        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            logging.error("No PDF files found in the input folder.")
            return

        start_time = time.time()
        total_files = len(pdf_paths)
        logging.info(f"Total PDF files to process: {total_files}")

        try:
            common_paragraphs, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")
            return

        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_folder, f"rationalized_result_{format_time}.xlsx")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Common Paragraphs"
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(common_paragraphs):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
            sheet.append([f"Paragraph {i + 1}", clean_paragraph])

        new_sheet = workbook.create_sheet(title="Matrix")
        header_row = ["PDF"] + [f"Paragraph {i + 1}" for i in range(len(common_paragraphs))]
        new_sheet.append(header_row)
        for row in matrix:
            new_sheet.append(row)

        workbook.save(output_file)
        workbook.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        logging.info(f"Processing completed in {elapsed_time:.2f} seconds.")
        logging.info(f"Results are saved in the file: {output_file}")

    def compare_similarity(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            logging.error("Input and output folders must be selected.")
            return

        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            logging.error("No PDF files found in the input folder.")
            return

        start_time = time.time()
        total_files = len(pdf_paths)
        logging.info(f"Total PDF files to process: {total_files}")

        try:
            all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")
            return

        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_folder, f"percentage_report_{format_time}.xlsx")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Paragraphs"
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(all_paragraphs):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
            sheet.append([f"Paragraph {i + 1}", clean_paragraph])

        new_sheet = workbook.create_sheet(title="Similarity Matrix")
        header_row = ["Paragraph"] + [f"Paragraph {i + 1}" for i in range(len(all_paragraphs))]
        new_sheet.append(header_row)
        matrix = []

        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        sorted_paragraphs = [sort_words(x) for x in all_paragraphs]

        for para1 in range(0, len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(0, len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()
                temp_list.append(round(s * 100, 2))
            matrix.append(temp_list)

        for row in range(len(matrix)):
            final_row = [f"Paragraph {row + 1}"] + matrix[row]
            new_sheet.append(final_row)

        workbook.save(output_file)
        workbook.close()

        end_time = time.time()
        elapsed_time = end_time - start_time
        logging.info(f"Processing completed in {elapsed_time:.2f} seconds.")
        logging.info(f"Results are saved in the file: {output_file}")

if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#1a1a2e")
    app = PDFComparerApp(root)
    root.mainloop()

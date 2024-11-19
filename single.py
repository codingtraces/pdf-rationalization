import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import fitz
import re
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from PIL import Image, ImageTk
import string
import threading
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from difflib import SequenceMatcher
from collections import defaultdict

if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

image1 = os.path.join(base_path, 'dev-logo.png')


class OneToManyComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("One to Many PDF Comparer Tool")

        # Variables to store input and output file paths
        self.single_pdf_folder_path = tk.StringVar()
        self.all_pdf_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for UI

        # Get the screen width and height
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        # Calculate the x and y positions to center the window
        x_position = (screen_width - 900) // 2
        y_position = (screen_height - 500) // 2

        # Set the window geometry
        self.master.geometry(f"900x500+{x_position}+{y_position}")

        # Heading
        heading_label = tk.Label(self.master, text="One to Many Content Comparer", font=("Helvetica", 26, "bold"), bg='#1a1a2e',
                                 fg='white')
        heading_label.pack(pady=30)

        # Single PDF Folder Location
        single_pdf_frame = tk.Frame(self.master)
        single_pdf_frame.pack(pady=15)

        single_pdf_label = tk.Label(single_pdf_frame, text="Single PDF Folder")
        single_pdf_label.pack(side=tk.LEFT)

        single_pdf_entry = tk.Entry(single_pdf_frame, textvariable=self.single_pdf_folder_path, width=50)
        single_pdf_entry.pack(side=tk.LEFT)

        single_pdf_button = tk.Button(single_pdf_frame, text="Browse", command=self.browse_single_pdf_folder)
        single_pdf_button.pack(side=tk.LEFT, padx=(10, 0))

        # All PDF Folder Location
        all_pdf_frame = tk.Frame(self.master)
        all_pdf_frame.pack(pady=15)

        all_pdf_label = tk.Label(all_pdf_frame, text="All PDFs Folder")
        all_pdf_label.pack(side=tk.LEFT)

        all_pdf_entry = tk.Entry(all_pdf_frame, textvariable=self.all_pdf_folder_path, width=50)
        all_pdf_entry.pack(side=tk.LEFT)

        all_pdf_button = tk.Button(all_pdf_frame, text="Browse", command=self.browse_all_pdf_folder)
        all_pdf_button.pack(side=tk.LEFT, padx=(10, 0))

        # Output File Location
        output_frame = tk.Frame(self.master)
        output_frame.pack(pady=15)

        output_label = tk.Label(output_frame, text="Output Folder")
        output_label.pack(side=tk.LEFT)

        output_entry = tk.Entry(output_frame, textvariable=self.output_folder_path, width=50)
        output_entry.pack(side=tk.LEFT)

        output_button = tk.Button(output_frame, text="Browse", command=self.browse_output_folder, bg="white")
        output_button.pack(side=tk.LEFT, padx=(10, 0))

        # Compare Buttons
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=20)

        compare_excel_button = tk.Button(compare_frame, text="Generate Excel Report", font=("Helvetica", 10, "bold"),
                                         command=lambda: threading.Thread(target=self.compare_one_to_many_excel).start(), width=20, height=2,
                                         bg='white')
        compare_excel_button.pack(side=tk.LEFT, padx=5)

        compare_html_button = tk.Button(compare_frame, text="Generate HTML Report", font=("Helvetica", 10, "bold"),
                                        command=lambda: threading.Thread(target=self.compare_one_to_many_html).start(), width=20, height=2,
                                        bg='white')
        compare_html_button.pack(side=tk.LEFT, padx=5)

        similarity_button = tk.Button(compare_frame, text="Generate Percentage Match", font=("Helvetica", 10, "bold"),
                                      command=lambda: threading.Thread(target=self.compare_similarity).start(), width=20, height=2, bg="white")
        similarity_button.pack(side=tk.LEFT, padx=5)

        original_image = Image.open(image1)
        self.photo = ImageTk.PhotoImage(original_image)
        self.image_label = tk.Label(self.master, image=self.photo)
        self.image_label.pack(side="right", padx=35)

    def browse_single_pdf_folder(self):
        folder_path = filedialog.askdirectory()
        self.single_pdf_folder_path.set(folder_path)

    def browse_all_pdf_folder(self):
        folder_path = filedialog.askdirectory()
        self.all_pdf_folder_path.set(folder_path)

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        self.output_folder_path.set(folder_path)

    def show_progress(self, message):
        messagebox.showinfo("Progress", message)

    def extract_paragraphs_from_pdf(self, file_path):
        paragraphs = []
        try:
            with fitz.open(file_path) as pdf_document:
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    text = page.get_text("text")
                    standardized_text = re.sub(r'\r\n|\r|\n', '\n', text)
                    page_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', standardized_text)
                    paragraphs.extend(page_paragraphs)

                    table_pattern = re.compile(r'\n[=]+\n')
                    table_matches = table_pattern.finditer(text)
                    for match in table_matches:
                        table_start = match.start()
                        table_end = match.end()

                        if table_end < len(text):
                            additional_paragraphs = re.split(r'[.!?](?:(?!\n)\s*\n\s*)', text[table_end:].strip())
                            paragraphs.extend(additional_paragraphs)

                    paragraphs = [p for p in paragraphs if p.strip() != ""]

        except Exception as e:
            messagebox.showerror("Error", f"Error extracting text from {file_path}: {str(e)}")

        return paragraphs

    def remove_illegal_characters(self, value):
        if isinstance(value, str):
            return ''.join(c for c in value if c in string.printable and c not in '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F')
        return value

    def compare_one_to_many_excel(self):
        try:
            self.show_progress("One to Many comparison for Excel report in progress...")
            # Call the comparison logic and generate an Excel report
            self.compare_one_to_many(report_type="excel")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def compare_one_to_many_html(self):
        try:
            self.show_progress("One to Many comparison for HTML report in progress...")
            # Call the comparison logic and generate an HTML report
            self.compare_one_to_many(report_type="html")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def compare_similarity(self):
        try:
            self.show_progress("Percentage match analysis in progress...")

            single_pdf_folder = self.single_pdf_folder_path.get()
            all_pdf_folder = self.all_pdf_folder_path.get()
            output_folder = self.output_folder_path.get()

            single_pdf_paths = [os.path.join(single_pdf_folder, f) for f in os.listdir(single_pdf_folder) if f.endswith('.pdf')]
            all_pdf_paths = [os.path.join(all_pdf_folder, f) for f in os.listdir(all_pdf_folder) if f.endswith('.pdf')]

            if not single_pdf_paths or not all_pdf_paths:
                messagebox.showerror("Error", "Please ensure both single and all PDFs folders contain PDF files.")
                return

            single_pdf_path = single_pdf_paths[0]
            single_paragraphs = self.extract_paragraphs_from_pdf(single_pdf_path)
            all_paragraphs = []

            for pdf_path in all_pdf_paths:
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                all_paragraphs.extend(paragraphs)

            sorted_single_paragraphs = [self.sort_words(x) for x in single_paragraphs]
            sorted_all_paragraphs = [self.sort_words(x) for x in all_paragraphs]

            similarity_scores = []
            for para1 in sorted_single_paragraphs:
                temp_scores = []
                for para2 in sorted_all_paragraphs:
                    m = SequenceMatcher(None, para1, para2)
                    temp_scores.append(round(m.ratio() * 100, 2))
                similarity_scores.append(temp_scores)

            # Create and save the percentage match report
            self.generate_percentage_match_report(single_paragraphs, similarity_scores, output_folder)
            self.show_progress("Percentage match analysis completed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def sort_words(self, paragraph):
        return ' '.join(sorted(paragraph.split()))

    def compare_one_to_many(self, report_type):
        try:
            self.show_progress(f"One to Many comparison ({report_type}) in progress...")

            single_pdf_folder = self.single_pdf_folder_path.get()
            all_pdf_folder = self.all_pdf_folder_path.get()
            output_folder = self.output_folder_path.get()

            single_pdf_paths = [os.path.join(single_pdf_folder, f) for f in os.listdir(single_pdf_folder) if f.endswith('.pdf')]
            all_pdf_paths = [os.path.join(all_pdf_folder, f) for f in os.listdir(all_pdf_folder) if f.endswith('.pdf')]

            if not single_pdf_paths or not all_pdf_paths:
                messagebox.showerror("Error", "Please ensure both single and all PDFs folders contain PDF files.")
                return

            # Assuming one single PDF to compare against multiple PDFs
            single_pdf_path = single_pdf_paths[0]
            single_paragraphs = self.extract_paragraphs_from_pdf(single_pdf_path)
            pdf_reports = {}
            common_elements = {"text_blocks": defaultdict(list)}

            for pdf_path in all_pdf_paths:
                paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
                pdf_reports[os.path.basename(pdf_path)] = paragraphs

            all_text_blocks = single_paragraphs + [block for report in pdf_reports.values() for block in report]
            pdf_names = ["Single PDF"] * len(single_paragraphs) + [pdf_name for pdf_name, report in pdf_reports.items() for _ in report]

            vectorizer = TfidfVectorizer().fit_transform(all_text_blocks)
            similarity_matrix = cosine_similarity(vectorizer)

            for i in range(len(single_paragraphs)):
                for j in range(len(single_paragraphs), len(all_text_blocks)):
                    similarity = similarity_matrix[i, j]
                    if similarity > 0.1:
                        common_elements["text_blocks"][single_paragraphs[i]].append((pdf_names[j], round(similarity * 100, 2)))

            if report_type == "excel":
                self.generate_comparison_excel_report(common_elements, output_folder)
            elif report_type == "html":
                self.generate_comparison_html_report(common_elements, output_folder)

            self.show_progress(f"One to Many comparison ({report_type}) completed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def generate_comparison_excel_report(self, common_elements, output_folder):
        try:
            rows = []
            for item, matches in common_elements["text_blocks"].items():
                for match in matches:
                    pdf_name, similarity = match
                    rows.append(["Text Block", item, pdf_name, similarity])

            df = pd.DataFrame(rows, columns=["Type", "Content (Paragraph)", "Found in PDF", "Similarity Percentage"])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template Reusability Report"

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                    if r_idx > 1:
                        if c_idx == 4:
                            similarity = float(value)
                            fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") if similarity == 100 else PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            ws.cell(row=r_idx, column=2).fill = fill
                            ws.cell(row=r_idx, column=4).fill = fill

            # Adding headers to the Excel file
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="A9D0F5", end_color="A9D0F5", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            current_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
            excel_filename = os.path.join(output_folder, f"template_reusability_report_{current_time}.xlsx")
            wb.save(excel_filename)
        except Exception as e:
            messagebox.showerror("Error", f"Error generating Excel report: {str(e)}")

    def generate_comparison_html_report(self, common_elements, output_folder):
        try:
            html_content = """
            <html>
            <head><title>Template Reusability Report</title></head>
            <body>
            <h1>Template Reusability Report</h1>
            <table border="1">
            <tr><th>Type</th><th>Content (Paragraph)</th><th>Found in PDF</th><th>Similarity Percentage</th></tr>
            """

            for item, matches in common_elements["text_blocks"].items():
                for match in matches:
                    pdf_name, similarity = match
                    html_content += f"<tr><td>Text Block</td><td>{item}</td><td>{pdf_name}</td><td>{similarity}%</td></tr>"

            html_content += """</table></body></html>"""

            current_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
            html_filename = os.path.join(output_folder, f"template_reusability_report_{current_time}.html")

            with open(html_filename, "w", encoding="utf-8") as file:
                file.write(html_content)
        except Exception as e:
            messagebox.showerror("Error", f"Error generating HTML report: {str(e)}")

    def generate_percentage_match_report(self, single_paragraphs, similarity_scores, output_folder):
        try:
            rows = []
            for i, scores in enumerate(similarity_scores):
                for j, score in enumerate(scores):
                    rows.append([f"Paragraph {i + 1}", f"Compared Paragraph {j + 1}", score])

            df = pd.DataFrame(rows, columns=["Source Paragraph", "Compared Paragraph", "Similarity Percentage"])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Percentage Match Report"

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Adding headers to the Excel file
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="A9D0F5", end_color="A9D0F5", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            current_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
            excel_filename = os.path.join(output_folder, f"percentage_match_report_{current_time}.xlsx")
            wb.save(excel_filename)
        except Exception as e:
            messagebox.showerror("Error", f"Error generating Percentage Match report: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg='#1a1a2e')
    app = OneToManyComparerApp(root)
    root.mainloop()

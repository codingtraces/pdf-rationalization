import os
import tkinter as tk
import sys
import fitz  # PyMuPDF, used for PDF operations
from tkinter import filedialog
import re
import openpyxl
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageTk
import threading
import logging
import time
import csv
import hashlib
from functools import lru_cache

# Configure logging for detailed debugging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Determine the base path for resources
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS  # If the script is compiled, use the temporary directory
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # Otherwise, use the script directory

# Paths to the logo images
image1 = os.path.join(base_path, 'dev-logo.png')
image2 = os.path.join(base_path, 'dev-logo.png')

def hash_paragraph(paragraph):
    return hashlib.sha256(paragraph.encode('utf-8')).hexdigest()

@lru_cache(maxsize=1024)
def extract_paragraphs_from_pdf_cached(file_path, file_modified_time):
    paragraphs = []
    try:
        logging.info(f"Extracting text from {file_path} using PyMuPDF.")
        with fitz.open(file_path) as doc:
            text = ""
            for page in doc:
                text += page.get_text()

        lines = text.splitlines()
        paragraph = ""
        for line in lines:
            if line.strip():
                if paragraph:
                    paragraph += " " + line.strip()
                else:
                    paragraph = line.strip()
            else:
                if paragraph:
                    paragraphs.append(paragraph.strip())
                    paragraph = ""
        if paragraph:
            paragraphs.append(paragraph.strip())

        combined_paragraphs = []
        temp_paragraph = ""
        for para in paragraphs:
            if len(para.split()) < 20:
                temp_paragraph += " " + para
            else:
                if temp_paragraph:
                    combined_paragraphs.append(temp_paragraph.strip())
                    temp_paragraph = ""
                combined_paragraphs.append(para)
        if temp_paragraph:
            combined_paragraphs.append(temp_paragraph.strip())

        return combined_paragraphs
    except Exception as e:
        logging.error(f"Error extracting text from {file_path}: {str(e)}")
    return paragraphs

class PDFComparerApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF Comparer Tool")

        # Variables to store input and output folder paths
        self.input_folder_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()

        # Create GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Tkinter widgets for the UI
        self.configure_window()
        self.create_heading_frame()
        self.create_input_output_frames()
        self.create_compare_buttons()

    def configure_window(self):
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x_position = (screen_width - 980) // 2
        y_position = (screen_height - 500) // 2
        self.master.geometry(f"980x500+{x_position}+{y_position}")

    def create_heading_frame(self):
        heading_frame = tk.Frame(self.master, bg="#1a1a2e")
        heading_frame.pack(fill=tk.X, pady=10, padx=10)

        self.load_image(heading_frame, image1, "left")
        heading_label = tk.Label(heading_frame, text="Content Rationalizer", font=("Helvetica", 26, "bold"),
                                 bg="#1a1a2e", fg="white")
        heading_label.pack(side="left", expand=True)
        self.load_image(heading_frame, image2, "right")

    def load_image(self, frame, image_path, side):
        try:
            if os.path.exists(image_path):
                original_image = Image.open(image_path).resize((100, 100), Image.LANCZOS)
                photo = ImageTk.PhotoImage(original_image)
                image_label = tk.Label(frame, image=photo, bg="#1a1a2e")
                image_label.image = photo  # Keep reference to avoid garbage collection
                image_label.pack(side=side, padx=10)
            else:
                raise FileNotFoundError(f"Image file not found: {image_path}")
        except Exception as e:
            logging.error(f"Error loading image: {str(e)}")

    def create_input_output_frames(self):
        self.create_folder_frame("Input Folder ", self.input_folder_path, self.browse_input_folder)
        self.create_folder_frame("Output Folder ", self.output_folder_path, self.browse_output_folder)

    def create_folder_frame(self, label_text, path_variable, browse_command):
        frame = tk.Frame(self.master)
        frame.pack(pady=20)
        label = tk.Label(frame, text=label_text)
        label.pack(side=tk.LEFT)
        entry = tk.Entry(frame, textvariable=path_variable, width=50)
        entry.pack(side=tk.LEFT)
        button = tk.Button(frame, text="Browse", command=browse_command)
        button.pack(side=tk.LEFT, padx=(10, 0))

    def create_compare_buttons(self):
        compare_frame = tk.Frame(self.master, bg="#1a1a2e")
        compare_frame.pack(pady=10)
        self.create_compare_button(compare_frame, "Rationalise (Excel)", self.compare_pdfs)
        self.create_compare_button(compare_frame, "Rationalise (CSV)", self.compare_pdfs_csv, padx=(15, 0))
        self.create_compare_button(compare_frame, "Percentage Match (Excel)", self.compare_similarity, padx=(15, 0))
        self.create_compare_button(compare_frame, "Percentage Match (CSV)", self.compare_similarity_csv, padx=(15, 0))

    def create_compare_button(self, frame, text, command, padx=None):
        button = tk.Button(frame, text=text, font=("Helvetica", 10, "bold"),
                           command=lambda: threading.Thread(target=command).start(), width=20,
                           height=2, bg="white")
        button.pack(side=tk.LEFT, padx=padx)

    def browse_input_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.input_folder_path.set(folder_path)

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)

    def extract_paragraphs_from_pdf(self, file_path):
        file_modified_time = os.path.getmtime(file_path)
        return extract_paragraphs_from_pdf_cached(file_path, file_modified_time)

    def compare_paragraphs(self, pdf_paths, compare):
        if compare == "pdfcompare":
            return self.perform_pdf_comparison(pdf_paths)
        elif compare == "percentage_match":
            return self.perform_percentage_match(pdf_paths)

    def perform_pdf_comparison(self, pdf_paths):
        all_hashes = set()
        pdf_hashes = {}
        for index, pdf_path in enumerate(pdf_paths):
            logging.info(f"Extracting paragraphs from {pdf_path} ({index + 1}/{len(pdf_paths)})")
            paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
            all_hashes.update(paragraphs)
            pdf_hashes[pdf_path] = paragraphs

        all_hashes = sorted(list(all_hashes))
        matrix = [[os.path.basename(pdf_path)] + [1 if hash_value in pdf_hashes[pdf_path] else 0 for hash_value in all_hashes] for pdf_path in pdf_paths]
        return all_hashes, matrix

    def perform_percentage_match(self, pdf_paths):
        all_paragraphs = []
        for index, pdf_path in enumerate(pdf_paths):
            logging.info(f"Extracting paragraphs from {pdf_path} ({index + 1}/{len(pdf_paths)})")
            paragraphs = self.extract_paragraphs_from_pdf(pdf_path)
            all_paragraphs.extend(paragraphs)
        return all_paragraphs

    def compare_pdfs(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            common_hashes, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")
            self.save_results_excel(output_folder, "rationalized_result", common_hashes, matrix)
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_pdfs_csv(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            common_hashes, matrix = self.compare_paragraphs(pdf_paths, "pdfcompare")
            self.save_results_csv(output_folder, "rationalized_result", common_hashes, matrix)
        except Exception as e:
            logging.error(f"Error during comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_similarity(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")
            self.save_similarity_excel(output_folder, "percentage_report", all_paragraphs)
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")

        self.log_processing_time(start_time)

    def compare_similarity_csv(self):
        input_folder, output_folder, pdf_paths = self.get_input_output_paths()
        if not pdf_paths:
            return

        start_time = time.time()
        logging.info(f"Total PDF files to process: {len(pdf_paths)}")

        try:
            all_paragraphs = self.compare_paragraphs(pdf_paths, "percentage_match")
            self.save_similarity_csv(output_folder, "percentage_report", all_paragraphs)
        except Exception as e:
            logging.error(f"Error during similarity comparison: {str(e)}")

        self.log_processing_time(start_time)

    def get_input_output_paths(self):
        input_folder = self.input_folder_path.get()
        output_folder = self.output_folder_path.get()

        if not input_folder or not output_folder:
            logging.error("Input and output folders must be selected.")
            return None, None, None

        pdf_paths = [os.path.join(input_folder, f) for f in os.listdir(input_folder) if f.endswith('.pdf')]
        if not pdf_paths:
            logging.error("No PDF files found in the input folder.")
            return None, None, None

        return input_folder, output_folder, pdf_paths

    def save_results_excel(self, output_folder, filename_prefix, common_hashes, matrix):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.xlsx")

        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Common Paragraphs")
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(common_hashes):
            sheet.append([f"Paragraph {i + 1}", paragraph])

        new_sheet = workbook.create_sheet(title="Matrix")
        header_row = ["PDF"] + [f"Paragraph {i + 1}" for i in range(len(common_hashes))]
        new_sheet.append(header_row)
        for row in matrix:
            new_sheet.append(row)

        workbook.save(output_file)
        workbook.close()
        logging.info(f"Results are saved in the file: {output_file}")

    def save_results_csv(self, output_folder, filename_prefix, common_hashes, matrix):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_csv_common = os.path.join(output_folder, f"{filename_prefix}_common_{format_time}.csv")
        output_csv_matrix = os.path.join(output_folder, f"{filename_prefix}_matrix_{format_time}.csv")

        # Write Common Paragraphs to CSV
        with open(output_csv_common, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Paragraph ID", "Hash"])
            for i, hash_value in enumerate(common_hashes):
                writer.writerow([f"Paragraph {i + 1}", hash_value])

        # Write Matrix to CSV
        with open(output_csv_matrix, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            header_row = ["PDF"] + [f"Paragraph {i + 1}" for i in range(len(common_hashes))]
            writer.writerow(header_row)
            for row in matrix:
                writer.writerow(row)

        logging.info(f"CSV Results are saved in the files: {output_csv_common}, {output_csv_matrix}")

    def save_similarity_excel(self, output_folder, filename_prefix, all_paragraphs):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_folder, f"{filename_prefix}_{format_time}.xlsx")

        workbook = openpyxl.Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Paragraphs")
        sheet.append(["Paragraph ID", "Content"])
        for i, paragraph in enumerate(all_paragraphs):
            clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
            sheet.append([f"Paragraph {i + 1}", clean_paragraph])

        new_sheet = workbook.create_sheet(title="Similarity Matrix")
        header_row = ["Paragraph"] + [f"Paragraph {i + 1}" for i in range(len(all_paragraphs))]
        new_sheet.append(header_row)
        matrix = self.calculate_similarity_matrix(all_paragraphs)

        for row in range(len(matrix)):
            final_row = [f"Paragraph {row + 1}"] + matrix[row]
            new_sheet.append(final_row)

        workbook.save(output_file)
        workbook.close()
        logging.info(f"Results are saved in the file: {output_file}")

    def save_similarity_csv(self, output_folder, filename_prefix, all_paragraphs):
        current_time = datetime.now()
        format_time = current_time.strftime("%Y%m%d%H%M%S")
        output_csv_paragraphs = os.path.join(output_folder, f"{filename_prefix}_paragraphs_{format_time}.csv")
        output_csv_matrix = os.path.join(output_folder, f"{filename_prefix}_matrix_{format_time}.csv")

        # Write Paragraphs to CSV
        with open(output_csv_paragraphs, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Paragraph ID", "Content"])
            for i, paragraph in enumerate(all_paragraphs):
                clean_paragraph = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', paragraph)
                writer.writerow([f"Paragraph {i + 1}", clean_paragraph])

        # Write Similarity Matrix to CSV
        matrix = self.calculate_similarity_matrix(all_paragraphs)
        with open(output_csv_matrix, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            header_row = ["Paragraph"] + [f"Paragraph {i + 1}" for i in range(len(all_paragraphs))]
            writer.writerow(header_row)
            for row_idx, row in enumerate(matrix):
                writer.writerow([f"Paragraph {row_idx + 1}"] + row)

        logging.info(f"CSV Results are saved in the files: {output_csv_paragraphs}, {output_csv_matrix}")

    def calculate_similarity_matrix(self, paragraphs):
        def sort_words(paragraph):
            return ' '.join(sorted(paragraph.split()))

        sorted_paragraphs = [sort_words(x) for x in paragraphs]
        matrix = []
        for para1 in range(0, len(sorted_paragraphs)):
            temp_list = []
            for para2 in range(0, len(sorted_paragraphs)):
                m = SequenceMatcher(None, sorted_paragraphs[para1], sorted_paragraphs[para2])
                s = m.ratio()
                temp_list.append(round(s * 100, 2))
            matrix.append(temp_list)
        return matrix

    def log_processing_time(self, start_time):
        end_time = time.time()
        elapsed_time = end_time - start_time
        logging.info(f"Processing completed in {elapsed_time:.2f} seconds.")

if __name__ == "__main__":
    root = tk.Tk()
    root.configure(bg="#1a1a2e")
    app = PDFComparerApp(root)
    root.mainloop()